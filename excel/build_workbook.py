import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
GOLD = "B08D57"
LIGHT = "EAF1F8"
WHITE = "FFFFFF"

df = pd.read_csv("../data/order_to_cash_clean.csv", parse_dates=["OrderDate", "InvoiceDate", "PaymentDate"])

wb = Workbook()

# ---------------------------------------------------------------
# Sheet: Data
# ---------------------------------------------------------------
ws_data = wb.active
ws_data.title = "Data"
cols = ["OrderID", "OrderDate", "InvoiceDate", "PaymentDate", "CustomerSegment", "Amount",
        "PaymentTerms", "IsCreditNote", "OrderToInvoiceDays", "InvoiceToPaymentDays",
        "OrderToPaymentDays", "IsPaid", "IsInvoiced"]
ws_data.append(cols)
for c in ws_data[1]:
    c.font = Font(bold=True, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", fgColor=NAVY)
for _, row in df.iterrows():
    ws_data.append([row[c] if pd.notna(row[c]) else None for c in cols])
n_rows = len(df) + 1
date_cols = ["OrderDate", "InvoiceDate", "PaymentDate"]
for dc in date_cols:
    col_idx = cols.index(dc) + 1
    for r in range(2, n_rows + 1):
        ws_data.cell(row=r, column=col_idx).number_format = "YYYY-MM-DD"
for i in range(1, len(cols) + 1):
    ws_data.column_dimensions[get_column_letter(i)].width = 14

def colref(name):
    return get_column_letter(cols.index(name) + 1)

# ---------------------------------------------------------------
# Sheet: Segment_Summary (formula-driven)
# ---------------------------------------------------------------
ws_s = wb.create_sheet("Segment_Summary")
segments = ["SME", "Enterprise", "Government"]
ws_s.append(["Segment", "PaidOrders", "AvgOrderToInvoiceDays", "AvgInvoiceToPaymentDays",
             "AvgTotalCycleDays", "OpenOrders", "OutstandingAR"])
for c in ws_s[1]:
    c.font = Font(bold=True, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", fgColor=NAVY)

CS, ICN, IP, O2I, I2P, O2P, AMT = (colref("CustomerSegment"), colref("IsCreditNote"),
    colref("IsPaid"), colref("OrderToInvoiceDays"), colref("InvoiceToPaymentDays"),
    colref("OrderToPaymentDays"), colref("Amount"))

for i, seg in enumerate(segments):
    r = i + 2
    crit = f'Data!${CS}$2:${CS}${n_rows},A{r},Data!${ICN}$2:${ICN}${n_rows},FALSE,Data!${IP}$2:${IP}${n_rows},TRUE'
    ws_s.cell(row=r, column=1, value=seg)
    ws_s.cell(row=r, column=2, value=f'=COUNTIFS({crit})')
    ws_s.cell(row=r, column=3, value=f'=ROUND(AVERAGEIFS(Data!${O2I}$2:${O2I}${n_rows},{crit}),1)')
    ws_s.cell(row=r, column=4, value=f'=ROUND(AVERAGEIFS(Data!${I2P}$2:${I2P}${n_rows},{crit}),1)')
    ws_s.cell(row=r, column=5, value=f'=ROUND(AVERAGEIFS(Data!${O2P}$2:${O2P}${n_rows},{crit}),1)')
    crit_open = f'Data!${CS}$2:${CS}${n_rows},A{r},Data!${ICN}$2:${ICN}${n_rows},FALSE,Data!${IP}$2:${IP}${n_rows},FALSE'
    ws_s.cell(row=r, column=6, value=f'=COUNTIFS({crit_open})')
    ws_s.cell(row=r, column=7, value=f'=ROUND(SUMIFS(Data!${AMT}$2:${AMT}${n_rows},{crit_open}),0)')
last_s_row = len(segments) + 1
for i in range(1, 8):
    ws_s.column_dimensions[get_column_letter(i)].width = 20

# ---------------------------------------------------------------
# Sheet: Cash_Flow_Scenario
# ---------------------------------------------------------------
ws_c = wb.create_sheet("Cash_Flow_Scenario")
ws_c["A1"] = "SCENARIO: Reduce Government Order-to-Invoice Delay by 7 Days"
ws_c["A1"].font = Font(bold=True, size=13, color=NAVY, name="Arial")
labels = ["Total Government Revenue (period)", "Period Length (days)", "Avg Daily Government Revenue",
          "Days Reduction (scenario input)", "Working Capital Freed (€)"]
for i, lbl in enumerate(labels):
    ws_c.cell(row=3 + i, column=1, value=lbl).font = Font(bold=True, name="Arial")

crit_gov_paid = f'Data!${CS}$2:${CS}${n_rows},"Government",Data!${ICN}$2:${ICN}${n_rows},FALSE,Data!${IP}$2:${IP}${n_rows},TRUE'
OD = colref("OrderDate")
ws_c["B3"] = f'=ROUND(SUMIFS(Data!${AMT}$2:${AMT}${n_rows},{crit_gov_paid}),0)'
ws_c["B4"] = f'=ROUND(MAX(Data!${OD}$2:${OD}${n_rows})-MIN(Data!${OD}$2:${OD}${n_rows}),0)'
ws_c["B5"] = "=ROUND(B3/B4,0)"
ws_c["B6"] = 7
ws_c["B7"] = "=ROUND(B5*B6,0)"
for r in (3, 4, 5, 6, 7):
    ws_c.cell(row=r, column=2).font = Font(size=12, color=GOLD, bold=True, name="Arial")
ws_c["B7"].font = Font(size=15, color="E34948", bold=True, name="Arial")
ws_c.column_dimensions["A"].width = 40
ws_c.column_dimensions["B"].width = 22

# ---------------------------------------------------------------
# Sheet: Dashboard
# ---------------------------------------------------------------
ws_d = wb.create_sheet("Dashboard", 0)
ws_d.sheet_view.showGridLines = False
ws_d.merge_cells("B2:K2")
ws_d["B2"] = "ORDER-TO-CASH PROCESS DASHBOARD"
ws_d["B2"].font = Font(bold=True, size=19, color=NAVY, name="Arial")
ws_d.merge_cells("B3:K3")
ws_d["B3"] = "DSO & Cycle-Time Analysis — SME · Enterprise · Government"
ws_d["B3"].font = Font(italic=True, size=12, color=GOLD, name="Arial")

def kpi_card(ws, col, label, formula, fmt="#,##0"):
    col_letter = get_column_letter(col)
    ws.merge_cells(f"{col_letter}5:{get_column_letter(col+1)}5")
    ws[f"{col_letter}5"] = label
    ws[f"{col_letter}5"].font = Font(bold=True, color=WHITE, size=10, name="Arial")
    ws[f"{col_letter}5"].fill = PatternFill("solid", fgColor=NAVY)
    ws[f"{col_letter}5"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws.merge_cells(f"{col_letter}6:{get_column_letter(col+1)}7")
    cell = ws[f"{col_letter}6"]
    cell.value = formula
    cell.font = Font(bold=True, size=17, color=GOLD, name="Arial")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt
    for rr in (5, 6, 7):
        for cc in (col, col + 1):
            ws.cell(row=rr, column=cc).border = Border(*(Side(style="thin", color="CCCCCC"),) * 4)

dso_crit_num = f'(Data!${ICN}$2:${ICN}${n_rows}=FALSE)*(Data!${IP}$2:${IP}${n_rows}=TRUE)*Data!${AMT}$2:${AMT}${n_rows}*Data!${I2P}$2:${I2P}${n_rows}'
dso_crit_den = f'(Data!${ICN}$2:${ICN}${n_rows}=FALSE)*(Data!${IP}$2:${IP}${n_rows}=TRUE)*Data!${AMT}$2:${AMT}${n_rows}'
kpi_card(ws_d, 2, "COMPANY-WIDE DSO", f'=ROUND(SUMPRODUCT({dso_crit_num})/SUMPRODUCT({dso_crit_den}),1)', "0.0")
kpi_card(ws_d, 4, "GOVERNMENT DSO", "=Segment_Summary!D4", "0.0")
kpi_card(ws_d, 6, "GOV. ORDER-TO-INVOICE (days)", "=Segment_Summary!C4", "0.0")
kpi_card(ws_d, 8, "OUTSTANDING AR (Government)", "=Segment_Summary!G4", '#,##0" €"')
kpi_card(ws_d, 10, "WORKING CAPITAL FREED (7-day fix)", "=Cash_Flow_Scenario!B7", '#,##0" €"')

ws_d.row_dimensions[6].height = 20
ws_d.row_dimensions[7].height = 20

bar = BarChart()
bar.title = "Avg Cycle Time by Stage & Segment (days)"
bar.y_axis.title = "Days"
data = Reference(ws_s, min_col=3, max_col=4, min_row=1, max_row=last_s_row)
cats = Reference(ws_s, min_col=1, min_row=2, max_row=last_s_row)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.width, bar.height = 18, 9
ws_d.add_chart(bar, "B10")

bar2 = BarChart()
bar2.title = "Outstanding Accounts Receivable by Segment (€)"
bar2.y_axis.title = "€"
data2 = Reference(ws_s, min_col=7, min_row=1, max_row=last_s_row)
cats2 = Reference(ws_s, min_col=1, min_row=2, max_row=last_s_row)
bar2.add_data(data2, titles_from_data=True)
bar2.set_categories(cats2)
bar2.width, bar2.height = 18, 9
ws_d.add_chart(bar2, "B29")

for i in range(1, 12):
    ws_d.column_dimensions[get_column_letter(i)].width = 15
ws_d.page_setup.orientation = "landscape"
ws_d.page_setup.fitToWidth = 1
ws_d.page_setup.fitToHeight = 0
ws_d.sheet_properties.pageSetUpPr.fitToPage = True

wb.save("Order_to_Cash_Dashboard.xlsx")
print("saved")
